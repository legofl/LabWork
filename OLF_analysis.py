'''
Python script to analyze OLF data excel files
Meant to compare the classifications assigned to certain gene variants in 2 different excel spreadsheets
Author: Lauren Fogel
'''
from openpyxl import load_workbook

# opens the files
data_R = load_workbook('Categories_of_and_criteria_for_myocilin_variants_2.xlsx')
data_AM = load_workbook('AlphaMissense_OLF.xlsx')

# tells which sheet to pull from
ws_R = data_R["data"]
ws_AM = data_AM.active
#print(ws_R.sheetnames)


# finds variant matches and prints out categories so they are all in the same file
def match():
    print(f'Review : AlphaMissense')
    for i in range(2,ws_R.max_row+1):
        for j in range(2,ws_AM.max_row+1):
            
            # check for name matches
            if (str(ws_R.cell(row=i,column=1).value) == str(ws_AM.cell(row=j,column=2).value)):
                # check classifications
                # match
                ws_R.cell(row=i, column=5, value = ws_AM.cell(row=j,column=4).value)
                data_R.save('Categories_of_and_criteria_for_myocilin_variants_2.xlsx')
                print(f'{ws_R.cell(row=i,column=3).value} : {ws_AM.cell(row=j,column=4).value}')



# comparisons - more complicated version of match() above
# if they are the same - strings should be the same 
# if likely pathogenic - add a different condition since this is not a condition in AM
def compare():
    for i in range(2,ws_R.max_row+1):
        for j in range(2,ws_AM.max_row+1):
            
            # check for name matches
            if (str(ws_R.cell(row=i,column=1).value) == str(ws_AM.cell(row=j,column=2).value)):
                # check classifications
                # match
                if (str(ws_R.cell(row=i,column=3).value) == str(ws_AM.cell(row=j,column=4).value)):
                    ws_R.cell(row=i, column=5, value = "same")
                    data_R.save('Categories_of_and_criteria_for_myocilin_variants.xlsx')
                    print("same")
                # likely pathogenic - special case
                elif (str(ws_R.cell(row=i,column=3).value) == "likely pathogenic"):
                    ws_R.cell(row=i, column=5, value = "no match: " + ws_AM.cell(row=j,column=4).value)
                    data_R.save('Categories_of_and_criteria_for_myocilin_variants.xlsx')
                    print("Likely pathogenic: ", ws_AM.cell(row=j,column=4).value)
                # does not match
                elif (str(ws_R.cell(row=i,column=3).value) != str(ws_AM.cell(row=j,column=4).value)):
                    ws_R.cell(row=i, column=5, value = "different: " + ws_AM.cell(row=j,column=4).value)
                    data_R.save('Categories_of_and_criteria_for_myocilin_variants.xlsx')
                    print("different: ", ws_AM.cell(row=j,column=4).value)
                # catch any other condition
                else:
                    ws_R.cell(row=i, column=5, value = "NA")
                    data_R.save('Categories_of_and_criteria_for_myocilin_variants.xlsx')
                    print("NA")
        

# Run whichever function suits your intentions
match()
#compare()
